import shutil

import openpyxl
import os
import cpca

import datetime


def convert_filed(row_has,ws):
    ws.cell(i + row_has, 1, source_sh.cell(i + 1, 8).value)
    dict = {'招标公告': '公告', '招标预告': '预告', '招标结果': '中标'}
    key = source_sh.cell(i + 1, 6).value
    ws.cell(i + row_has, 2, dict[key] if key in dict else key)
    ws.cell(i + row_has, 3, source_sh.cell(i + 1, 7).value)
    str = [source_sh.cell(i + 1, 5).value]
    df = cpca.transform(str)
    ws.cell(i + row_has, 7, df.iloc[0, 0])
    ws.cell(i + row_has, 8, df.iloc[0, 1])
    ws.cell(i + row_has, 9, source_sh.cell(i + 1, 9).value)
    ws.cell(i + row_has, 11, title)
    ws.cell(i + row_has, 12, '')
    ws.cell(i + row_has, 13, '')
    ws.cell(i + row_has, 14, '')
    dttm = datetime.datetime.strptime(source_sh.cell(i + 1, 2).value, "%Y-%m-%d")
    ws.cell(i + row_has, 15, dttm)
    ws.cell(i + row_has, 15).number_format = 'yyyy/m/d;@'
    segment = ''
    if (title.find('公司') != -1):
        segment = 'A'
    elif (title.find('医院') != -1 or title.find('卫生院')):
        segment = 'P'
    elif (title.find('大学') != -1):
        segment = 'R'
    elif (title.find('研究所') != -1):
        segment = 'S'
    ws.cell(i + row_has, 16, segment)
    if content.find('附件') != -1:
        ws.cell(i + row_has, 17, 'true')
    ws.cell(i + row_has, 18, hp_source_sh.cell(16, 2).value)
    link = hp_source_sh.cell(16, 2).value
    ws.cell(i + row_has, 18, '= HYPERLINK("{}","{}")'.format(link, link))
    ws.cell(i + row_has, 19, content)


if __name__ == '__main__':
    sheet_name = '基本信息'
    # 创建导出表格
    if (not os.path.exists('target.xlsx')):
        shutil.copy('template.xlsx', 'target.xlsx')
    wb = openpyxl.load_workbook('target.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    ws_drop = wb.get_sheet_by_name('过滤掉')
    # 读取
    for root, dir, file in os.walk('sourceExcel'):
        for f in file:
            source_wb = openpyxl.load_workbook('sourceExcel/' + f)
            # 选择表单
            source_sh = source_wb.get_sheet_by_name(sheet_name)
            row_has = ws.max_row
            row_has_drop=ws_drop.max_row
            for i in range(1, source_sh.max_row - 2):
                # 处理超链接中的正文
                hp = source_sh.cell(i + 1, 1).hyperlink
                location = hp.location
                start = location.find('#') + 1
                end = location.find('!')
                hp_sheet_name = location[start:end]
                hp_source_sh = source_wb.get_sheet_by_name(hp_sheet_name)
                # 抽取字段
                title = source_sh.cell(i + 1, 1).value
                content = hp_source_sh.cell(17, 2).value
                # 过滤
                if (content.find('耗材') != -1 or title.find('耗材') != -1):
                    row_has-=1
                    convert_filed(row_has_drop,ws_drop)
                else:
                # 转换字段
                    row_has_drop-=1
                    convert_filed(row_has,ws)
    #删除空行
    # empty_rows = []
    # for idx, row in enumerate(ws.iter_rows(50), 1):
    #     empty = not any((cell.value for cell in row))
    #     if empty:
    #         empty_rows.append(idx)
    # for row_idx in reversed(empty_rows):
    #     ws.delete_rows(row_idx, 1)
    #保存
    wb.save("target.xlsx")

